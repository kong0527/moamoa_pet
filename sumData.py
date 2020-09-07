# !/usr/bin/python3
import openpyxl
import sys
import csv


def convert_excel(csvFile):
    # convert from csv to xlsx
    wb = openpyxl.Workbook()
    ws = wb.active

    with open(csvFile) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)

    fileName = csvFile.replace(".csv", ".xlsx")
    wb.save(fileName)


if __name__ == "__main__":
    inputFile1 = sys.argv[1]
    inputFile2 = sys.argv[2]

    # If you have both xlsx files, skip this sentence
    ###########
    if inputFile1 != 'total.xlsx':
        convert_excel(inputFile1)
        inputFile1 = inputFile1.replace(".csv", ".xlsx")

    convert_excel(inputFile2)
    inputFile2 = inputFile2.replace(".csv", ".xlsx")
    ############

    wb1 = openpyxl.load_workbook(inputFile1)
    ws1 = wb1.active

    wb2 = openpyxl.load_workbook(inputFile2)
    ws2 = wb2.active

    wb = openpyxl.Workbook()
    ws = wb.active
    i = 2
    while True:
        if not ws2.cell(row=i, column=1).value:
            break
        ws2.cell(row=i, column=3, value="{}".format(0))
        i = i + 1

    i = 2
    while True:
        if not ws1.cell(row=i, column=1).value:
            break
        A_name = ws1.cell(row=i, column=1).value
        # eng ver
        # A_name = A_name.lower()
        # A_name = A_name.capitalize()
        A_count = int(ws1.cell(row=i, column=2).value)

        ws.cell(row=i, column=1, value="{}".format(A_name))
        ws.cell(row=i, column=2, value="{}".format(A_count))
        j = 2
        while True:
            if not ws2.cell(row=j, column=1).value:
                break

            B_name = ws2.cell(row=j, column=1).value
            # eng ver
            # B_name = B_name.lower()
            # B_name = B_name.capitalize()
            B_count = int(ws2.cell(row=j, column=2).value)

            if A_name == B_name:
                total_count = A_count + B_count
                ws.cell(row=i, column=2, value="{}".format(total_count))
                ws2.cell(row=i, column=3, value="{}".format(1))
            j = j + 1
        if ws.cell(row=j, column=2).value == 0:
            ws.cell(row=i, column=2, value="{}".format(A_count))
        i = i + 1

    k = 2
    while True:
        if not ws2.cell(row=k, column=1).value:
            break
        B_name = ws2.cell(row=k, column=1).value
        # eng ver
        # B_name = B_name.lower()
        # B_name = B_name.capitalize()
        B_count = int(ws2.cell(row=k, column=2).value)

        if ws2.cell(row=k, column=3).value == 0:
            ws.cell(row=i, column=1, value="{}".format(B_name))
            ws.cell(row=i, column=2, value="{}".format(B_count))
        i = i + 1
        k = k + 1

    wb.save('total.csv')
    wb.save('total.xlsx')
